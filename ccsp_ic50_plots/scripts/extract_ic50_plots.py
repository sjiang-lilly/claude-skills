#!/usr/bin/env python3
"""
Extract IC50 % Inhibition dose-response plots from CCSP Excel files and generate an HTML summary table.

Author: Mandy Jiang (shan.jiang2@lilly.com)
Organization: Eli Lilly and Company - Bioinformatics
Version: 1.3.0
Created: February 2026

Usage:
    python extract_ic50_plots.py <input> <output_html> [--compound-map compound_map.json] [--cell-colors cell_colors.json]

Arguments:
    input: Path to zip file or folder containing CCSP Excel files (*_paste.xlsx)
    output_html: Output HTML file path

Optional:
    --compound-map: JSON file mapping compound IDs to display names
    --cell-colors: JSON file mapping cell lines to background colors

Key Features:
    - Extracts % Inhibition plots ONLY (not Response plots)
    - Dynamic compound detection from XLFit Chart section (matches plot order)
    - Staurosporine (control) is EXCLUDED from extraction and output
    - Number of plots = number of test compounds in Excel
"""

import os
import sys
import re
import json
import base64
import zipfile
import shutil
import subprocess
import argparse
import tempfile
import warnings
from pathlib import Path

warnings.filterwarnings('ignore')


def get_cell_line(filename):
    """Extract cell line name from filename.
    
    Handles format: YYYYMMDD_CELLLINE_NTA_NNH_paste.xlsx → CELLLINE
    """
    parts = filename.split('_')
    return parts[1] if len(parts) >= 2 else filename.replace('.xlsx', '')


def is_staurosporine(compound_name):
    """Check if compound is Staurosporine (control)."""
    if not compound_name:
        return False
    name_lower = compound_name.lower()
    return 'staurosporine' in name_lower or 'stauro' in name_lower


def extract_compounds_from_excel(excel_path):
    """Extract test compound IDs from XLFit Chart section in Excel file.
    
    The XLFit Chart section defines the actual plot order, which matches
    the embedded EMF images. Staurosporine is excluded.
    
    Returns:
        list: Test compound IDs in plot order (Staurosporine excluded)
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, data_only=True)
        
        # Try multiple possible sheet names
        sheet_names_to_try = ["Analyzed Data", "Data analysis for IC50"]
        sheet = None
        for name in sheet_names_to_try:
            if name in wb.sheetnames:
                sheet = wb[name]
                break
        
        if sheet is None:
            wb.close()
            return []
        
        compounds = []
        in_xlfit_section = False
        
        # Find XLFit Chart section and extract compounds in plot order
        for row in range(1, 50):
            col_b = sheet.cell(row=row, column=2).value
            
            if col_b and isinstance(col_b, str):
                # Found XLFit Chart marker - start extracting
                if 'XLFit' in col_b and 'Chart' in col_b:
                    in_xlfit_section = True
                    continue
                
                if in_xlfit_section:
                    compound_id = _extract_compound_id(col_b)
                    # Exclude Staurosporine (control)
                    if compound_id and not is_staurosporine(compound_id):
                        compounds.append(compound_id)
            
            # Stop if we hit empty row after finding compounds
            elif in_xlfit_section and not col_b and len(compounds) > 0:
                break
        
        wb.close()
        return compounds
        
    except Exception as e:
        print(f"  Warning: Could not extract compounds: {e}", file=sys.stderr)
        return []


def _extract_compound_id(cell_value):
    """Extract compound ID from cell value.

    Handles formats:
        - CELLLINE_COMPOUND → COMPOUND (e.g., BT20_TA145 → TA145)
        - COMPOUND → COMPOUND
    """
    if not cell_value or not isinstance(cell_value, str):
        return None

    cell_value = cell_value.strip()
    if '_' in cell_value:
        parts = cell_value.split('_')
        # Return second part (compound ID), not the cell line prefix
        return parts[1] if len(parts) >= 2 else cell_value
    return cell_value


def img_to_base64(img_path):
    """Convert image file to base64 string for HTML embedding."""
    with open(img_path, "rb") as f:
        return base64.b64encode(f.read()).decode('utf-8')


def extract_and_convert_images(excel_path, output_dir, cell_line, num_compounds):
    """Extract EMF images from Excel and convert to PNG.
    
    Only extracts the first N images (where N = num_compounds) which are
    the % Inhibition plots for test compounds. Staurosporine plot and 
    Response plots are excluded.
    
    Args:
        excel_path: Path to Excel file
        output_dir: Directory to save converted images
        cell_line: Cell line name for organizing output
        num_compounds: Number of test compounds (excluding Staurosporine)
    
    Returns:
        list: Filenames of converted PNG images in order
    """
    cell_output_dir = os.path.join(output_dir, cell_line)
    os.makedirs(cell_output_dir, exist_ok=True)
    
    temp_dir = os.path.join(output_dir, f".temp_{cell_line}")
    os.makedirs(temp_dir, exist_ok=True)
    
    try:
        # Extract xlsx (it's a ZIP archive)
        with zipfile.ZipFile(excel_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        media_dir = os.path.join(temp_dir, "xl", "media")
        if not os.path.exists(media_dir):
            return []
        
        emf_files = [f for f in os.listdir(media_dir) if f.endswith('.emf')]
        
        # Sort by image number
        def get_num(fname):
            match = re.search(r'image(\d+)\.emf', fname)
            return int(match.group(1)) if match else 0
        
        emf_files_sorted = sorted(emf_files, key=get_num)
        
        # Filter to actual plots (exclude small placeholder images)
        large_emfs = []
        for emf_file in emf_files_sorted:
            emf_path = os.path.join(media_dir, emf_file)
            if os.path.getsize(emf_path) > 3000:  # Actual plots are > 3KB
                large_emfs.append(emf_file)
        
        # Take first N images ONLY where N = number of test compounds
        # This gets % Inhibition plots for test compounds only
        # Excludes: Staurosporine plot and all Response plots
        emf_to_convert = large_emfs[:num_compounds]
        
        converted = []
        for emf_file in emf_to_convert:
            emf_path = os.path.join(media_dir, emf_file)
            png_name = emf_file.replace('.emf', '.png')
            png_path = os.path.join(cell_output_dir, png_name)
            
            try:
                result = subprocess.run(
                    ['inkscape', emf_path, '--export-filename', png_path],
                    capture_output=True, text=True, timeout=60
                )
                if os.path.exists(png_path) and os.path.getsize(png_path) > 1000:
                    converted.append(png_name)
            except subprocess.TimeoutExpired:
                print(f"  Warning: Timeout converting {emf_file}", file=sys.stderr)
            except Exception as e:
                print(f"  Warning: Could not convert {emf_file}: {e}", file=sys.stderr)
        
        return converted
        
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def generate_html(cell_compound_plots, output_path, compounds, compound_names=None, cell_colors=None):
    """Generate HTML table with IC50 % Inhibition plots.
    
    Args:
        cell_compound_plots: Dict mapping cell_line -> {compound: image_path}
        output_path: Output HTML file path
        compounds: List of test compound IDs (Staurosporine excluded)
        compound_names: Optional dict mapping compound IDs to display names
        cell_colors: Optional dict mapping cell lines to background colors
    """
    compound_names = compound_names or {}
    cell_colors = cell_colors or {}
    
    html_content = """<!DOCTYPE html>
<html>
<head>
    <title>IC50 Plots - % Inhibition</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; background: #fff; }
        h1 { color: #333; text-align: center; margin-bottom: 5px; }
        h2 { color: #666; text-align: center; font-weight: normal; margin-top: 0; }
        table { border-collapse: collapse; margin: 20px auto; background: white; }
        th, td { border: 1px solid #ccc; padding: 5px; text-align: center; vertical-align: middle; }
        th { background-color: #2E7D32; color: white; font-size: 11px; padding: 8px 5px; max-width: 120px; }
        th .compound-id { font-weight: bold; display: block; }
        th .compound-name { font-weight: normal; font-size: 10px; display: block; margin-top: 2px; }
        th.cell-line-header { background-color: #1565C0; min-width: 80px; }
        td.cell-line { font-weight: bold; font-size: 12px; }
        td img { width: 160px; height: auto; display: block; }
        .caption { text-align: center; color: #666; font-size: 12px; margin-top: 10px; }
    </style>
</head>
<body>
    <h1>IC50 Dose-Response Curves (% Inhibition)</h1>
    <h2>CCSP Screening Data</h2>
    <table>
        <tr>
            <th class="cell-line-header">Cell Line</th>
"""
    
    # Add compound headers (test compounds only, Staurosporine excluded)
    for compound in compounds:
        display_name = compound_names.get(compound, '')
        if display_name:
            html_content += f'            <th><span class="compound-id">{compound}</span><span class="compound-name">{display_name}</span></th>\n'
        else:
            html_content += f'            <th><span class="compound-id">{compound}</span></th>\n'
    html_content += '        </tr>\n'
    
    # Add rows for each cell line (sorted alphabetically)
    for cell_line in sorted(cell_compound_plots.keys()):
        bg_color = cell_colors.get(cell_line, '#FFFFFF')
        html_content += f'        <tr>\n            <td class="cell-line" style="background-color: {bg_color};">{cell_line}</td>\n'
        
        for compound in compounds:
            img_path = cell_compound_plots[cell_line].get(compound)
            if img_path and os.path.exists(img_path):
                b64_data = img_to_base64(img_path)
                html_content += f'            <td><img src="data:image/png;base64,{b64_data}" alt="{compound}_{cell_line}"></td>\n'
            else:
                html_content += '            <td>-</td>\n'
        
        html_content += '        </tr>\n'
    
    num_cells = len(cell_compound_plots)
    num_compounds = len(compounds)
    html_content += f"""    </table>
    <p class="caption">{num_compounds} test compounds × {num_cells} cell lines (% Inhibition plots only)</p>
</body>
</html>
"""
    
    with open(output_path, 'w') as f:
        f.write(html_content)


def find_all_excel_files(input_path):
    """Find all Excel files from input path, handling zip extraction and multiple subfolders.

    Args:
        input_path: Path to zip file or folder

    Returns:
        tuple: (list of (excel_file_path, cell_line) tuples, temp_dir_to_cleanup_or_None)
    """
    excel_files_found = []

    def is_ccsp_file(filename):
        """Check if file matches CCSP pattern: *_paste.xlsx"""
        return (filename.endswith('_paste.xlsx') and
                not filename.startswith('~') and
                'Summary' not in filename)

    # If it's a zip file, extract it
    if input_path.endswith('.zip') and os.path.isfile(input_path):
        temp_dir = tempfile.mkdtemp(prefix='ccsp_extract_')
        print(f"Extracting zip file to temporary directory...")
        with zipfile.ZipFile(input_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)

        # Find all Excel files (may be nested in subfolders)
        for root, dirs, files in os.walk(temp_dir):
            # Skip __MACOSX folders
            if '__MACOSX' in root:
                continue
            for f in files:
                if is_ccsp_file(f):
                    file_path = os.path.join(root, f)
                    cell_line = get_cell_line(f)
                    excel_files_found.append((file_path, cell_line))

        return excel_files_found, temp_dir

    # If it's a directory, search it and all subfolders
    elif os.path.isdir(input_path):
        # Check files directly in this folder
        for f in os.listdir(input_path):
            file_path = os.path.join(input_path, f)
            if os.path.isfile(file_path) and is_ccsp_file(f):
                cell_line = get_cell_line(f)
                excel_files_found.append((file_path, cell_line))

        # Check all subfolders
        for item in os.listdir(input_path):
            subfolder = os.path.join(input_path, item)
            if os.path.isdir(subfolder) and not item.startswith('__') and not item.startswith('.'):
                for f in os.listdir(subfolder):
                    file_path = os.path.join(subfolder, f)
                    if os.path.isfile(file_path) and is_ccsp_file(f):
                        cell_line = get_cell_line(f)
                        excel_files_found.append((file_path, cell_line))

        return excel_files_found, None

    else:
        print(f"Error: Input path '{input_path}' is not a valid zip file or directory", file=sys.stderr)
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description='Extract IC50 % Inhibition plots from CCSP Excel files',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python extract_ic50_plots.py data.zip output.html
  python extract_ic50_plots.py /path/to/folder output.html
  python extract_ic50_plots.py data.zip output.html --compound-map names.json
        """
    )
    parser.add_argument('input', help='Zip file or folder containing CCSP Excel files')
    parser.add_argument('output_html', help='Output HTML file path')
    parser.add_argument('--compound-map', help='JSON file with compound ID to display name mapping')
    parser.add_argument('--cell-colors', help='JSON file with cell line to color mapping')
    args = parser.parse_args()
    
    # Load optional compound mapping
    compound_names = {}
    if args.compound_map and os.path.exists(args.compound_map):
        with open(args.compound_map) as f:
            compound_names = json.load(f)
        print(f"Loaded compound name mapping: {len(compound_names)} entries")
    
    # Load optional cell colors
    cell_colors = {}
    if args.cell_colors and os.path.exists(args.cell_colors):
        with open(args.cell_colors) as f:
            cell_colors = json.load(f)
        print(f"Loaded cell color mapping: {len(cell_colors)} entries")
    
    # Find all Excel files (handles zip extraction and multiple subfolders)
    excel_files, temp_extract_dir = find_all_excel_files(args.input)

    try:
        if not excel_files:
            print("Error: No CCSP Excel files found", file=sys.stderr)
            sys.exit(1)

        # Sort by cell line name
        excel_files = sorted(excel_files, key=lambda x: x[1])

        print(f"Found {len(excel_files)} cell line files")

        # Extract test compounds from first file (excludes Staurosporine)
        first_file = excel_files[0][0]
        compounds = extract_compounds_from_excel(first_file)

        if not compounds:
            print("Error: Could not extract compound list from Excel files", file=sys.stderr)
            sys.exit(1)

        print(f"Found {len(compounds)} test compounds: {', '.join(compounds)}")
        print(f"(Staurosporine excluded)")

        # Create temp directory for images
        temp_images_dir = os.path.join(os.path.dirname(args.output_html) or '.', '.ic50_plots_temp')
        os.makedirs(temp_images_dir, exist_ok=True)

        # Process each Excel file
        cell_compound_plots = {}

        for file_path, cell_line in excel_files:
            print(f"Processing {cell_line}...")

            # Extract only % Inhibition plots for test compounds
            converted = extract_and_convert_images(file_path, temp_images_dir, cell_line, len(compounds))

            if converted:
                cell_dir = os.path.join(temp_images_dir, cell_line)

                # Sort converted images by number
                def get_num(fname):
                    match = re.search(r'image(\d+)\.png', fname)
                    return int(match.group(1)) if match else 0
                converted_sorted = sorted(converted, key=get_num)

                # Map images to compounds (1:1 mapping in order)
                cell_compound_plots[cell_line] = {}
                for i, img in enumerate(converted_sorted):
                    if i < len(compounds):
                        cell_compound_plots[cell_line][compounds[i]] = os.path.join(cell_dir, img)

                print(f"  Extracted {len(converted_sorted)} % Inhibition plots")
        
        # Generate HTML output
        generate_html(cell_compound_plots, args.output_html, compounds, compound_names, cell_colors)
        
        print(f"\n✓ Output saved to: {args.output_html}")
        print(f"  {len(cell_compound_plots)} cell lines × {len(compounds)} test compounds")
        
        # Cleanup temp images
        shutil.rmtree(temp_images_dir, ignore_errors=True)
        
    finally:
        # Cleanup temp extraction directory if we created one
        if temp_extract_dir:
            shutil.rmtree(temp_extract_dir, ignore_errors=True)


if __name__ == '__main__':
    main()
