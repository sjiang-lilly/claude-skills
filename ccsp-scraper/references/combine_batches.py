import pandas as pd
import pyreadr
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

# =============================================================================
# USER CONFIGURATION
# =============================================================================

# Select gene markers to include (choose from available markers below)
# Available: TACSTD2, CEACAM5, ERBB2, ERBB3, MET, EGFR, CD276, F3, MUC1, PTK7,
#            ITGB6, FOLR1, DLL3, ADAM9, LRRC15, FAP, ITGAV, AXL, CDCP1, TPBG,
#            CEACAM6, CLDN18, MSLN, MUC16, CDH17, SLFN11, TOP1, ABCB1, ABCC3,
#            ABCG2, PAF1
SELECTED_GENE_MARKERS = ['TACSTD2', 'CEACAM5']

# =============================================================================
# SCRIPT START
# =============================================================================

# Base directory
base_dir = '/Users/L052239/Library/CloudStorage/OneDrive-EliLillyandCompany/Desktop/Bioinformatics/Projects/ClaudeCode/CCSP_Scraper'

# =============================================================================
# STEP 1: Load compound mapping
# =============================================================================

print("Loading compound mapping...")
compound_list_file = os.path.join(base_dir, 'CompoundList.xlsx')
df_compounds = pd.read_excel(compound_list_file)
# Create mapping: TA_ID -> Compound name
compound_mapping = dict(zip(df_compounds['TA_ID'], df_compounds['Compound']))
print(f"Compound mapping: {compound_mapping}\n")

# Find all batch folders (ending with "cells")
batch_folders = [d for d in os.listdir(base_dir) if d.endswith('cells') and os.path.isdir(os.path.join(base_dir, d))]
batch_folders.sort()

print(f"Found {len(batch_folders)} batch folders: {batch_folders}\n")

# =============================================================================
# STEP 2-3: Read assay data from each batch
# =============================================================================

def process_numeric_value(val):
    """Process numeric values: handle '>' prefix and round to 2 decimal places."""
    if pd.isna(val):
        return val
    val_str = str(val)
    # If value contains '>', extract the numeric part
    if '>' in val_str:
        val_str = val_str.replace('>', '').strip()
    try:
        return round(float(val_str), 2)
    except ValueError:
        return val

all_data = []

for batch_folder in batch_folders:
    batch_path = os.path.join(base_dir, batch_folder)
    excel_files = glob.glob(os.path.join(batch_path, '*.xlsx'))

    print(f"Processing batch: {batch_folder} ({len(excel_files)} files)")

    for file in excel_files:
        filename = os.path.basename(file)
        # Extract cell line name from filename (e.g., "20260122_BXPC3_6TA_144H_paste.xlsx" -> "BXPC3")
        cell_line = filename.split('_')[1]

        # Read Summary sheet
        df = pd.read_excel(file, sheet_name='Summary', header=1)

        # Select target columns
        df = df[['Compound ID', 'Max % Inhibition', 'Corrected Abs IC50 nM']]

        # Remove cell line prefix from Compound ID (e.g., "BxPC3_TA145" -> "TA145")
        df['Compound ID'] = df['Compound ID'].apply(lambda x: '_'.join(str(x).split('_')[1:]))

        # Map TA IDs to compound names (keep original if not in mapping, e.g., "Staurosporine")
        df['Compound ID'] = df['Compound ID'].apply(lambda x: compound_mapping.get(x, x))

        # Process numeric values: handle '>' and round to 2 decimal places
        df['Max % Inhibition'] = df['Max % Inhibition'].apply(process_numeric_value)
        df['Corrected Abs IC50 nM'] = df['Corrected Abs IC50 nM'].apply(process_numeric_value)

        # Add cell line and batch information
        df['Cell Line'] = cell_line
        df['Batch'] = batch_folder

        all_data.append(df)
        print(f"  Read {cell_line}: {len(df)} compounds")

# =============================================================================
# STEP 3: Pivot to wide format
# =============================================================================

# Combine all data
combined = pd.concat(all_data, ignore_index=True)
print(f"\nTotal records: {len(combined)}")

# Pivot for IC50 values (column names without suffix)
ic50_pivot = combined.pivot_table(
    index=['Cell Line', 'Batch'],
    columns='Compound ID',
    values='Corrected Abs IC50 nM',
    aggfunc='first'
)

# Pivot for Max % Inhibition values (column names without suffix)
inhib_pivot = combined.pivot_table(
    index=['Cell Line', 'Batch'],
    columns='Compound ID',
    values='Max % Inhibition',
    aggfunc='first'
)

# Get compound names (sorted)
compound_names = sorted(ic50_pivot.columns.tolist())

# Combine pivots horizontally
pivot_combined = pd.concat([ic50_pivot[compound_names], inhib_pivot[compound_names]], axis=1)
pivot_combined = pivot_combined.reset_index()

# Rename columns to avoid duplicates (temporary, for merging)
ic50_cols = [f"{c}_IC50" for c in compound_names]
inhib_cols = [f"{c}_INHIB" for c in compound_names]
pivot_combined.columns = ['Cell Line', 'Batch'] + ic50_cols + inhib_cols

# =============================================================================
# STEP 4: Load metadata from RData
# =============================================================================

print("\nLoading metadata...")
rdata_file = os.path.join(base_dir, 'samplemeta_TAA_DepMap25Q2.RData')
result = pyreadr.read_r(rdata_file)
df_meta = result['df_merge']

# Cell line info columns
cell_info_cols = ['ModelID', 'CellLine', 'OncotreeLineage', 'OncotreePrimaryDisease',
                  'OncotreeSubtype', 'OncotreeCode', 'in_BioMetas']

# Validate selected gene markers
available_markers = ['TACSTD2', 'CEACAM5', 'ERBB2', 'ERBB3', 'MET', 'EGFR', 'CD276', 'F3',
                     'MUC1', 'PTK7', 'ITGB6', 'FOLR1', 'DLL3', 'ADAM9', 'LRRC15', 'FAP',
                     'ITGAV', 'AXL', 'CDCP1', 'TPBG', 'CEACAM6', 'CLDN18', 'MSLN', 'MUC16',
                     'CDH17', 'SLFN11', 'TOP1', 'ABCB1', 'ABCC3', 'ABCG2', 'PAF1']

for marker in SELECTED_GENE_MARKERS:
    if marker not in available_markers:
        print(f"WARNING: Gene marker '{marker}' not found in metadata. Skipping.")
        SELECTED_GENE_MARKERS.remove(marker)

print(f"Selected gene markers: {SELECTED_GENE_MARKERS}")

# Create metadata subset
meta_cols = cell_info_cols + SELECTED_GENE_MARKERS
df_meta_subset = df_meta[meta_cols].copy()

# Round gene marker values to 2 decimal places
for marker in SELECTED_GENE_MARKERS:
    df_meta_subset[marker] = df_meta_subset[marker].apply(lambda x: round(x, 2) if pd.notna(x) else x)

# =============================================================================
# STEP 5: Merge pivot table with metadata (left join)
# =============================================================================

# Create uppercase version for matching
df_meta_subset['CellLine_upper'] = df_meta_subset['CellLine'].str.upper()
pivot_combined['CellLine_upper'] = pivot_combined['Cell Line'].str.upper()

# Left join to keep all assay cell lines
final = pivot_combined.merge(df_meta_subset, on='CellLine_upper', how='left')

# Drop temporary columns and reorder
final = final.drop(columns=['CellLine_upper', 'Cell Line'])

# Reorder columns: CellLine, Batch, Metadata, Gene markers, IC50, Max%Inhib
metadata_cols = ['CellLine', 'Batch', 'ModelID', 'OncotreeLineage', 'OncotreePrimaryDisease',
                 'OncotreeSubtype', 'OncotreeCode', 'in_BioMetas']
final = final[metadata_cols + SELECTED_GENE_MARKERS + ic50_cols + inhib_cols]

# Rename IC50 and INHIB columns back to compound names only
rename_dict = {f"{c}_IC50": c for c in compound_names}
rename_dict.update({f"{c}_INHIB": c for c in compound_names})
# Can't rename to same name, so we'll handle this in Excel writing

# Sort by Batch then CellLine
final = final.sort_values(['Batch', 'CellLine']).reset_index(drop=True)

# =============================================================================
# STEP 6: Create Excel with merged header row
# =============================================================================

output_file = os.path.join(base_dir, 'combined_all_batches.xlsx')

# Prepare column names (remove _IC50 and _INHIB suffixes)
final_columns = metadata_cols + SELECTED_GENE_MARKERS + compound_names + compound_names
final.columns = final_columns

# Write data to Excel first
final.to_excel(output_file, index=False, header=True, startrow=1)

# Now open and modify: add header indicator row and merge cells
wb = load_workbook(output_file)
ws = wb.active

# Write header indicator values in row 1
# Columns 1-2: blank (CellLine, Batch)
ws.cell(row=1, column=1).value = ''
ws.cell(row=1, column=2).value = ''

# Columns 3-8: Metadata (6 columns)
col_idx = 3
ws.cell(row=1, column=col_idx).value = 'Metadata'
metadata_end = col_idx + 6 - 1
ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=metadata_end)
ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
ws.cell(row=1, column=col_idx).font = Font(bold=True)
col_idx = metadata_end + 1

# Gene columns
ws.cell(row=1, column=col_idx).value = 'Gene'
gene_end = col_idx + len(SELECTED_GENE_MARKERS) - 1
if gene_end >= col_idx:
    ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=gene_end)
    ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
    ws.cell(row=1, column=col_idx).font = Font(bold=True)
col_idx = gene_end + 1

# IC50 columns
ws.cell(row=1, column=col_idx).value = 'IC50'
ic50_end = col_idx + len(compound_names) - 1
if ic50_end >= col_idx:
    ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=ic50_end)
    ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
    ws.cell(row=1, column=col_idx).font = Font(bold=True)
col_idx = ic50_end + 1

# Max%Inhib columns
ws.cell(row=1, column=col_idx).value = 'Max%Inhib'
inhib_end = col_idx + len(compound_names) - 1
if inhib_end >= col_idx:
    ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=inhib_end)
    ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
    ws.cell(row=1, column=col_idx).font = Font(bold=True)

wb.save(output_file)

print(f"\nOutput saved to: {output_file}")
print(f"Shape: {final.shape[0]} rows x {final.shape[1]} columns")
print(f"\nColumn breakdown:")
print(f"  - CellLine, Batch: 2 columns")
print(f"  - Metadata: 6 columns")
print(f"  - Gene markers: {len(SELECTED_GENE_MARKERS)} columns ({', '.join(SELECTED_GENE_MARKERS)})")
print(f"  - IC50: {len(compound_names)} columns")
print(f"  - Max%Inhib: {len(compound_names)} columns")
print(f"\nCell lines per batch:")
print(final.groupby('Batch')['CellLine'].count())
